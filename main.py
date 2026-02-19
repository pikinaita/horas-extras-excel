#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Registro de horas extraordinarias preservando ABSOLUTAMENTE TODO
el formato del Excel original: imágenes, estilos, dimensiones,
celdas combinadas, filas ocultas, bordes, fuentes, etc.

Solo se modifican los VALORES de celdas específicas.
"""

import os
import json
from datetime import datetime
from typing import Optional
from shutil import copy2
import tempfile

import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    import openpyxl
except ImportError:
    import sys, subprocess
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

# ── Constantes ────────────────────────────────────────────────────────────────

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
CONFIG_FILE = os.path.join(BASE_DIR, "config.json")

MESES_ES = {
    1: "Enero", 2: "Febrero", 3: "Marzo",
    4: "Abril", 5: "Mayo", 6: "Junio",
    7: "Julio", 8: "Agosto", 9: "Septiembre",
    10: "Octubre", 11: "Noviembre", 12: "Diciembre",
}

# Columnas de turno (1-based para openpyxl: C=3, D=4, E=5)
SHIFT_COLUMNS = {"Mañana": 3, "Tarde": 4, "Noche": 5}
COVERAGE_OPTIONS = ["Sustitución Operador", "Sustitución Jefe de Turno"]
JUSTIF_COL = 16  # Columna P
APP_VERSION = "2.1.0"

# ── Clase principal ───────────────────────────────────────────────────────────

class HorasExtrasApp:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title("Horas Extraordinarias - Parte Mensual")
        self.root.resizable(False, False)
        self._center(520, 580)
        self._apply_styles()

        self.config = self._load_config()
        self.filepath = self.config.get("filepath")

        self.container = ttk.Frame(self.root, padding=22)
        self.container.pack(fill=tk.BOTH, expand=True)

        if not self.filepath or not os.path.isfile(self.filepath):
            self._step_file()
        else:
            self._step_form()

    def _center(self, w: int, h: int) -> None:
        self.root.update_idletasks()
        x = (self.root.winfo_screenwidth() - w) // 2
        y = (self.root.winfo_screenheight() - h) // 2
        self.root.geometry(f"{w}x{h}+{x}+{y}")

    def _apply_styles(self) -> None:
        s = ttk.Style(self.root)
        try:
            s.theme_use("clam")
        except tk.TclError:
            pass
        s.configure("Title.TLabel", font=("Segoe UI", 13, "bold"))
        s.configure("Sub.TLabel", font=("Segoe UI", 9), foreground="#666677")
        s.configure("OK.TLabel", font=("Segoe UI", 9), foreground="#15803d")
        s.configure("Big.TButton", font=("Segoe UI", 10, "bold"), padding=6)

    def _clear(self) -> None:
        for w in self.container.winfo_children():
            w.destroy()

    def _load_config(self) -> dict:
        if os.path.isfile(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception:
                pass
        return {}

    def _save_config(self) -> None:
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump({"filepath": self.filepath}, f, ensure_ascii=False, indent=2)
        except Exception as e:
            messagebox.showwarning("Advertencia", "No se pudo guardar la configuración. " + str(e))

    def _step_file(self) -> None:
        self._clear()
        ttk.Label(self.container, text="Seleccionar archivo de partes", style="Title.TLabel").pack(anchor="w", pady=(0, 3))
        ttk.Label(self.container, text="Indique la ruta del libro Excel con los partes mensuales.", style="Sub.TLabel").pack(anchor="w", pady=(0, 14))
        ttk.Separator(self.container).pack(fill=tk.X, pady=(0, 16))

        ttk.Label(self.container, text="Ruta del archivo (.xlsx):").pack(anchor="w", pady=(0, 4))
        row = ttk.Frame(self.container)
        row.pack(fill=tk.X, pady=(0, 20))

        self._path_var = tk.StringVar(value=self.filepath or "")
        ttk.Entry(row, textvariable=self._path_var, width=52).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 6))
        ttk.Button(row, text="Examinar…", command=self._browse).pack(side=tk.LEFT)

        ttk.Button(self.container, text="Continuar →", style="Big.TButton", command=self._confirm_file).pack(pady=(4, 0))

    def _browse(self) -> None:
        path = filedialog.askopenfilename(
            title="Seleccionar parte de trabajo",
            filetypes=[("Libros Excel", "*.xlsx *.xlsm"), ("Todos los archivos", "*.*")]
        )
        if path:
            self._path_var.set(path)

    def _confirm_file(self) -> None:
        path = self._path_var.get().strip()
        if not path:
            messagebox.showerror("Error", "Debe seleccionar un archivo.")
            return
        if not os.path.isfile(path):
            messagebox.showerror("Error", "No se encontró el archivo indicado. Verifique la ruta.")
            return
        self.filepath = path
        self._save_config()
        self._step_form()

    def _step_form(self) -> None:
        self._clear()
        header = ttk.Frame(self.container)
        header.pack(fill=tk.X, pady=(0, 2))
        ttk.Label(header, text="Registrar horas extraordinarias", style="Title.TLabel").pack(side=tk.LEFT)
        ttk.Button(header, text="Cambiar archivo", command=self._step_file).pack(side=tk.RIGHT)

        fname = os.path.basename(self.filepath or "")
        ttk.Label(self.container, text=f"Archivo: {fname}", style="Sub.TLabel").pack(anchor="w", pady=(0, 10))
        ttk.Separator(self.container).pack(fill=tk.X, pady=(0, 12))

        # 1. Fecha
        frame_fecha = ttk.LabelFrame(self.container, text=" 1. Fecha de las horas extra ", padding=12)
        frame_fecha.pack(fill=tk.X, pady=(0, 10))
        row1 = ttk.Frame(frame_fecha)
        row1.pack()

        ttk.Label(row1, text="Día:").grid(row=0, column=0, sticky="e", padx=(0, 6))
        self._day_var = tk.StringVar(value=str(datetime.now().day))
        ttk.Spinbox(row1, from_=1, to=31, textvariable=self._day_var, width=5, font=("Segoe UI", 10)).grid(row=0, column=1, padx=(0, 20))

        ttk.Label(row1, text="Mes:").grid(row=0, column=2, sticky="e", padx=(0, 6))
        self._month_var = tk.StringVar()
        cb = ttk.Combobox(row1, textvariable=self._month_var, values=list(MESES_ES.values()), state="readonly", width=14)
        cb.set(MESES_ES[datetime.now().month])
        cb.grid(row=0, column=3)

        # 2. Turno
        frame_turno = ttk.LabelFrame(self.container, text=" 2. Turno ", padding=12)
        frame_turno.pack(fill=tk.X, pady=(0, 10))
        row2 = ttk.Frame(frame_turno)
        row2.pack()
        self._shift_var = tk.StringVar(value="Mañana")
        for col, nombre in enumerate(["Mañana", "Tarde", "Noche"]):
            ttk.Radiobutton(row2, text=nombre, variable=self._shift_var, value=nombre).grid(row=0, column=col, padx=26)

        # 3. Cobertura
        frame_cov = ttk.LabelFrame(self.container, text=" 3. ¿A quién cubre? ", padding=12)
        frame_cov.pack(fill=tk.X, pady=(0, 16))
        row3 = ttk.Frame(frame_cov)
        row3.pack()
        self._cov_var = tk.StringVar(value=COVERAGE_OPTIONS[0])
        for col, opt in enumerate(COVERAGE_OPTIONS):
            ttk.Radiobutton(row3, text=opt, variable=self._cov_var, value=opt).grid(row=0, column=col, padx=12, sticky="w")

        # Botones
        btns = ttk.Frame(self.container)
        btns.pack(fill=tk.X, pady=(4, 0))
        ttk.Button(btns, text="Registrar y guardar", style="Big.TButton", command=lambda: self._register(save_mode="inplace")).pack(side=tk.TOP, fill=tk.X, pady=(0, 6))
        ttk.Button(btns, text="Registrar en copia…", command=lambda: self._register(save_mode="saveas")).pack(side=tk.TOP, fill=tk.X)

        self._status_var = tk.StringVar()
        ttk.Label(self.container, textvariable=self._status_var, style="OK.TLabel").pack(anchor="w", pady=(10, 0))
        ttk.Label(self.container, text=f"v{APP_VERSION}", style="Sub.TLabel").pack(side=tk.BOTTOM, anchor="e")

    @staticmethod
    def _find_sheet(wb, month_name: str):
        mn = month_name.strip().lower()
        for name in wb.sheetnames:
            if name.strip().lower().startswith(mn):
                return wb[name]
        return None

    @staticmethod
    def _find_day_row(ws, day: int) -> Optional[int]:
        from datetime import datetime as _dt
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=1):
            cell = row[0]
            val = cell.value
            if val is None: continue
            if isinstance(val, str) and val.upper().startswith("=DAY("):
                date_val = ws.cell(row=cell.row, column=2).value
                if isinstance(date_val, _dt) and date_val.day == day:
                    return cell.row
            else:
                try:
                    if int(val) == day: return cell.row
                except (ValueError, TypeError): continue
        return None

    def _register(self, save_mode: str) -> None:
        try:
            day = int(self._day_var.get())
            if not 1 <= day <= 31: raise ValueError
        except Exception:
            messagebox.showerror("Error", "Introduzca un día válido (1-31).")
            return
        month = self._month_var.get()
        if not month:
            messagebox.showerror("Error", "Seleccione un mes.")
            return
        shift = self._shift_var.get()
        coverage = self._cov_var.get()
        try:
            with tempfile.TemporaryDirectory() as tmp_dir:
                temp_path = os.path.join(tmp_dir, "temp_partes.xlsx")
                copy2(self.filepath, temp_path)
                wb = openpyxl.load_workbook(temp_path, keep_vba=True)
                ws = self._find_sheet(wb, month)
                if ws is None:
                    messagebox.showerror("Hoja no encontrada", f"No existe hoja para '{month}'")
                    wb.close(); return
                row_idx = self._find_day_row(ws, day)
                if row_idx is None:
                    messagebox.showerror("Día no encontrado", f"No se encontró el día {day}")
                    wb.close(); return
                shift_cell = ws.cell(row=row_idx, column=SHIFT_COLUMNS[shift])
                justif_cell = ws.cell(row=row_idx, column=JUSTIF_COL)
                from copy import copy
                sf = copy(shift_cell.font); sb = copy(shift_cell.border)
                sfi = copy(shift_cell.fill); sa = copy(shift_cell.alignment)
                jf = copy(justif_cell.font); jb = copy(justif_cell.border)
                jfi = copy(justif_cell.fill); ja = copy(justif_cell.alignment)
                shift_cell.value = "X"
                justif_cell.value = coverage
                shift_cell.font, shift_cell.border, shift_cell.fill, shift_cell.alignment = sf, sb, sfi, sa
                justif_cell.font, justif_cell.border, justif_cell.fill, justif_cell.alignment = jf, jb, jfi, ja
                if save_mode == "inplace":
                    dest = self.filepath
                else:
                    dest = filedialog.asksaveasfilename(
                        title="Guardar copia del parte",
                        defaultextension=".xlsx",
                        initialfile=os.path.basename(self.filepath),
                        filetypes=[("Libros Excel", "*.xlsx *.xlsm"), ("Todos los archivos", "*.*")]
                    )
                    if not dest: wb.close(); return
                wb.save(dest)
                wb.close()
                self._status_var.set(f"Registrado: {day} de {month} ({shift})")
                messagebox.showinfo("Éxito", f"Datos guardados en:
{dest}")
        except PermissionError:
            messagebox.showerror("Archivo bloqueado", "Cierre el archivo Excel antes de registrar.")
        except Exception as e:
            messagebox.showerror("Error", str(e))

def main() -> None:
    root = tk.Tk()
    HorasExtrasApp(root)
    root.mainloop()

if __name__ == "__main__":
    main()
